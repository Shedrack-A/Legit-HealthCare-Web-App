from flask import Blueprint, jsonify, request, current_app, session, send_from_directory, send_file
from .models import User, Patient, ScreeningBioData, Consultation, FullBloodCount, KidneyFunctionTest, LipidProfile, LiverFunctionTest, ECG, Spirometry, Audiometry, Role, Permission, TemporaryAccessCode, AuditLog, SystemConfig, Conversation, Message, Branding
from . import db
import jwt
import os
from werkzeug.utils import secure_filename
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta, timezone, date
from functools import wraps
from sqlalchemy import func
import pyotp
import qrcode
import io
import base64
import smtplib
from email.message import EmailMessage

bp = Blueprint('api', __name__, url_prefix='/api')

@bp.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok'})

@bp.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No input data provided'}), 400

    username = data.get('username')
    email = data.get('email')
    password = data.get('password')

    if User.query.filter_by(username=username).first():
        return jsonify({'message': 'Username already exists'}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({'message': 'Email already registered'}), 400

    if not User.is_password_strong(password):
        return jsonify({'message': 'Password is not strong enough'}), 400

    new_user = User(
        first_name=data.get('first_name'),
        last_name=data.get('last_name'),
        other_name=data.get('other_name'),
        username=username,
        phone_number=data.get('phone_number'),
        email=email
    )
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    log_audit(new_user, 'USER_REGISTER', f"User {new_user.username} registered.")
    db.session.commit()

    return jsonify({'message': 'User registered successfully'}), 201

@bp.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    current_app.logger.info(f"Login attempt for user: {data.get('username')}")
    if not data or not data.get('username') or not data.get('password'):
        current_app.logger.warning("Login failed: No username or password provided.")
        return jsonify({'message': 'Could not verify'}), 401

    user = User.query.filter_by(username=data.get('username')).first()

    if not user:
        current_app.logger.warning(f"Login failed: User '{data.get('username')}' not found.")
        return jsonify({'message': 'Could not verify'}), 401

    if not user.check_password(data.get('password')):
        current_app.logger.warning(f"Login failed: Incorrect password for user '{data.get('username')}'.")
        return jsonify({'message': 'Could not verify'}), 401

    current_app.logger.info(f"Login successful for user: {user.username}")
    log_audit(user, 'USER_LOGIN', f"User {user.username} logged in.")
    db.session.commit()
    token = jwt.encode({
        'user_id': user.id,
        'exp': datetime.now(timezone.utc) + timedelta(minutes=30)
    }, current_app.config['SECRET_KEY'], algorithm="HS256")

    return jsonify({'token': token})

def token_required(func_or_permission=None):
    if callable(func_or_permission):
        permission = None
        f = func_or_permission
    else:
        permission = func_or_permission
        f = None

    def decorator(func):
        @wraps(func)
        def decorated_function(*args, **kwargs):
            token = None
            if 'Authorization' in request.headers:
                token = request.headers['Authorization'].split(" ")[1]

            if not token:
                return jsonify({'message': 'Token is missing!'}), 401

            try:
                data = jwt.decode(token, current_app.config['SECRET_KEY'], algorithms=["HS256"])
                current_user = User.query.get(data['user_id'])
            except:
                return jsonify({'message': 'Token is invalid!'}), 401

            if permission:
                user_permissions = {p.name for role in current_user.roles for p in role.permissions}
                if permission not in user_permissions:
                    temp_permissions = session.get('temp_permissions', {})
                    if permission not in temp_permissions or datetime.fromisoformat(temp_permissions[permission]) <= datetime.utcnow():
                        return jsonify({'message': 'You do not have permission to perform this action.'}), 403

            return func(current_user, *args, **kwargs)
        return decorated_function

    if f:
        return decorator(f)
    return decorator

@bp.route('/profile')
@token_required
def profile(current_user):
    user_data = {
        'id': current_user.id,
        'first_name': current_user.first_name,
        'last_name': current_user.last_name,
        'username': current_user.username,
        'email': current_user.email,
        'roles': [r.name for r in current_user.roles]
    }
    return jsonify(user_data)

@bp.route('/patients', methods=['POST'])
@token_required('register_patient')
def create_patient(current_user):
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No input data provided'}), 400

    required_fields = ['staff_id', 'first_name', 'last_name', 'department', 'gender', 'date_of_birth', 'contact_phone', 'email_address', 'race', 'nationality']
    if not all(field in data for field in required_fields):
        return jsonify({'message': 'Missing required fields'}), 400

    if Patient.query.filter_by(staff_id=data['staff_id']).first():
        return jsonify({'message': 'Patient with this Staff ID already exists'}), 400

    patient_id = f"P{datetime.now().year}{Patient.query.count() + 1:04d}"
    dob = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()

    new_patient = Patient(
        staff_id=data['staff_id'],
        patient_id=patient_id,
        first_name=data['first_name'],
        middle_name=data.get('middle_name'),
        last_name=data['last_name'],
        department=data['department'],
        gender=data['gender'],
        date_of_birth=dob,
        contact_phone=data['contact_phone'],
        email_address=data['email_address'],
        race=data['race'],
        nationality=data['nationality']
    )

    db.session.add(new_patient)
    log_audit(current_user, 'PATIENT_REGISTER', f"Registered patient {new_patient.first_name} {new_patient.last_name} (Staff ID: {new_patient.staff_id})")
    db.session.commit()

    return jsonify({'message': 'Patient registered successfully'}), 201

@bp.route('/patient-summary/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_patient_summary(current_user, staff_id):
    patient = Patient.query.filter_by(staff_id=staff_id).first_or_404()

    # Get the latest screening record for this patient to fetch year-specific data
    screening_record = ScreeningBioData.query.filter_by(
        patient_comprehensive_id=patient.id
    ).order_by(ScreeningBioData.screening_year.desc()).first()

    summary = {
        'patient_id': patient.id,
        'staff_id': patient.staff_id,
        'first_name': patient.first_name,
        'middle_name': patient.middle_name,
        'last_name': patient.last_name,
        'department': patient.department,
        'gender': patient.gender,
        'date_of_birth': patient.date_of_birth.isoformat(),
        'age': (date.today().year - patient.date_of_birth.year),
        'contact_phone': patient.contact_phone,
        'email_address': patient.email_address,
        'race': patient.race,
        'nationality': patient.nationality,
        'date_registered': screening_record.date_registered.isoformat() if screening_record else None,
        'patient_id_for_year': screening_record.patient_id_for_year if screening_record else None,
    }

    def model_to_dict(model_instance, exclude=None):
        if not model_instance:
            return {}
        if exclude is None:
            exclude = []

        result = {}
        for c in model_instance.__table__.columns:
            if c.name not in exclude:
                value = getattr(model_instance, c.name)
                if isinstance(value, (datetime, date)):
                    result[c.name] = value.isoformat()
                else:
                    result[c.name] = value
        return result

    summary.update(model_to_dict(patient.consultation, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.full_blood_count, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.kidney_function_test, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.lipid_profile, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.liver_function_test, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.ecg, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.spirometry, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.audiometry, exclude=['id', 'patient_id']))

    return jsonify(summary)

@bp.route('/me/report-summary', methods=['GET'])
@token_required
def get_my_report_summary(current_user):
    """
    Fetches the comprehensive report summary for the currently logged-in user,
    if they are linked to a patient profile.
    """
    patient = current_user.patient
    if not patient:
        return jsonify({'message': 'No patient profile linked to this user account.'}), 404

    # This logic is duplicated from get_patient_summary.
    # In a larger application, this could be refactored into a helper function.
    screening_record = ScreeningBioData.query.filter_by(
        patient_comprehensive_id=patient.id
    ).order_by(ScreeningBioData.screening_year.desc()).first()

    summary = {
        'patient_id': patient.id,
        'staff_id': patient.staff_id,
        'first_name': patient.first_name,
        'middle_name': patient.middle_name,
        'last_name': patient.last_name,
        'department': patient.department,
        'gender': patient.gender,
        'date_of_birth': patient.date_of_birth.isoformat(),
        'age': (date.today().year - patient.date_of_birth.year),
        'contact_phone': patient.contact_phone,
        'email_address': patient.email_address,
        'race': patient.race,
        'nationality': patient.nationality,
        'date_registered': screening_record.date_registered.isoformat() if screening_record else None,
        'patient_id_for_year': screening_record.patient_id_for_year if screening_record else None,
    }

    def model_to_dict(model_instance, exclude=None):
        if not model_instance:
            return {}
        if exclude is None:
            exclude = []

        result = {}
        for c in model_instance.__table__.columns:
            if c.name not in exclude:
                value = getattr(model_instance, c.name)
                if isinstance(value, (datetime, date)):
                    result[c.name] = value.isoformat()
                else:
                    result[c.name] = value
        return result

    summary.update(model_to_dict(patient.consultation, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.full_blood_count, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.kidney_function_test, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.lipid_profile, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.liver_function_test, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.ecg, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.spirometry, exclude=['id', 'patient_id']))
    summary.update(model_to_dict(patient.audiometry, exclude=['id', 'patient_id']))

    return jsonify(summary)

@bp.route('/save-director-review/<string:staff_id>', methods=['POST'])
@token_required('perform_director_review')
def save_director_review(current_user, staff_id):
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No input data provided'}), 400

    patient = Patient.query.filter_by(staff_id=staff_id).first_or_404()

    def update_model(model_class, patient_id, fields):
        obj = model_class.query.filter_by(patient_id=patient_id).first()
        if not obj:
            obj = model_class(patient_id=patient_id)
            db.session.add(obj)
        for field in fields:
            if field in data and hasattr(obj, field):
                setattr(obj, field, data[field])

    # Update Consultation
    consultation_fields = [
        'diabetes_mellitus', 'hypertension', 'bp', 'pulse', 'spo2', 'hs', 'breast_exam',
        'breast_exam_remark', 'abdomen', 'prostrate_specific_antigen', 'psa_remark',
        'fbs', 'rbs', 'fbs_rbs_remark', 'urine_analysis', 'ua_remark',
        'assessment_hx_pe', 'other_assessments', 'overall_lab_remark', 'other_remarks',
        'overall_assessment', 'comment_one', 'comment_two', 'comment_three', 'comment_four'
    ]
    update_model(Consultation, patient.id, consultation_fields)
    if not patient.consultation.director_review_timestamp:
        patient.consultation.director_review_timestamp = datetime.utcnow()

    # Update Test Results
    update_model(FullBloodCount, patient.id, ['hct', 'wbc', 'plt', 'lymp_percent', 'lymp', 'gra_percent', 'gra', 'mid_percent', 'mid', 'rbc', 'mcv', 'mch', 'mchc', 'rdw', 'pdw', 'hgb', 'fbc_remark', 'other_remarks'])
    update_model(KidneyFunctionTest, patient.id, ['k', 'na', 'cl', 'ca', 'hc03', 'urea', 'cre', 'kft_remark', 'other_remarks'])
    update_model(LipidProfile, patient.id, ['tcho', 'tg', 'hdl', 'ldl', 'lp_remark', 'other_remarks'])
    update_model(LiverFunctionTest, patient.id, ['ast', 'alt', 'alp', 'tb', 'cb', 'lft_remark', 'other_remarks'])
    update_model(ECG, patient.id, ['ecg_result', 'remark'])
    update_model(Spirometry, patient.id, ['spirometry_result', 'spirometry_remark'])
    update_model(Audiometry, patient.id, ['audiometry_result', 'audiometry_remark'])

    log_audit(current_user, 'DIRECTOR_REVIEW_SAVE', f"Saved director review for patient {patient.first_name} {patient.last_name} (Staff ID: {patient.staff_id})")
    db.session.commit()

    return jsonify({'message': 'Director review saved successfully.'}), 200

@bp.route('/patients', methods=['GET'])
@token_required('view_patient_data')
def get_all_patients(current_user):
    staff_id = request.args.get('staff_id')
    if staff_id:
        patient = Patient.query.filter_by(staff_id=staff_id).first()
        if not patient:
            return jsonify({'message': 'Patient not found'}), 404

        patient_data = {
            'staff_id': patient.staff_id,
            'patient_id': patient.patient_id,
            'first_name': patient.first_name,
            'middle_name': patient.middle_name,
            'last_name': patient.last_name,
            'department': patient.department,
            'gender': patient.gender,
            'date_of_birth': patient.date_of_birth.isoformat(),
            'contact_phone': patient.contact_phone,
            'email_address': patient.email_address,
            'race': patient.race,
            'nationality': patient.nationality,
        }
        return jsonify(patient_data)

    else:
        patients = Patient.query.order_by(Patient.first_name, Patient.last_name).all()
        results = [{
            'staff_id': p.staff_id,
            'first_name': p.first_name,
            'last_name': p.last_name,
            'department': p.department,
            'gender': p.gender,
            'contact_phone': p.contact_phone,
        } for p in patients]
        return jsonify(results)

@bp.route('/patient/<string:staff_id>', methods=['DELETE'])
@token_required('delete_patient')
def delete_patient(current_user, staff_id):
    patient = Patient.query.filter_by(staff_id=staff_id).first_or_404()
    log_audit(current_user, 'PATIENT_DELETE', f"Deleted patient {patient.first_name} {patient.last_name} (Staff ID: {staff_id})")
    db.session.delete(patient)
    db.session.commit()
    return jsonify({'message': 'Patient deleted successfully.'}), 200

@bp.route('/patient/<string:staff_id>', methods=['PUT'])
@token_required('edit_patient')
def update_patient(current_user, staff_id):
    patient = Patient.query.filter_by(staff_id=staff_id).first_or_404()
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No input data provided'}), 400

    updatable_fields = [
        'staff_id', 'first_name', 'middle_name', 'last_name', 'department',
        'gender', 'date_of_birth', 'contact_phone', 'email_address',
        'race', 'nationality'
    ]

    for field in updatable_fields:
        if field in data:
            if field == 'date_of_birth':
                setattr(patient, field, datetime.strptime(data[field], '%Y-%m-%d').date())
            else:
                setattr(patient, field, data[field])

    log_audit(current_user, 'PATIENT_UPDATE', f"Updated patient {patient.first_name} {patient.last_name} (Staff ID: {staff_id})")
    db.session.commit()
    return jsonify({'message': 'Patient updated successfully.'}), 200

@bp.route('/patients/claim-account', methods=['POST'])
def claim_account():
    data = request.get_json()
    staff_id = data.get('staff_id')
    email = data.get('email')
    password = data.get('password')

    if not all([staff_id, email, password]):
        return jsonify({'message': 'Staff ID, email, and password are required.'}), 400

    patient = Patient.query.filter_by(staff_id=staff_id, email_address=email).first()

    if not patient:
        return jsonify({'message': 'Invalid Staff ID or email address.'}), 404

    if patient.user_id:
        return jsonify({'message': 'This patient account has already been claimed.'}), 409

    if not User.is_password_strong(password):
        return jsonify({'message': 'Password is not strong enough.'}), 400

    username = f"patient_{staff_id}"
    if User.query.filter_by(username=username).first():
        return jsonify({'message': 'A user account for this patient already exists. Please contact support.'}), 409

    new_user = User(
        first_name=patient.first_name,
        last_name=patient.last_name,
        username=username,
        email=patient.email_address,
        phone_number=patient.contact_phone
    )
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()

    patient.user_id = new_user.id
    log_audit(new_user, 'PATIENT_ACCOUNT_CLAIM', f"Patient {patient.first_name} {patient.last_name} claimed account.")
    db.session.commit()

    return jsonify({'message': 'Account claimed successfully. You can now log in.'}), 201

@bp.route('/screening/records', methods=['GET'])
@token_required('view_screening_records')
def get_screening_records(current_user):
    screening_year = request.args.get('screening_year', type=int)
    company_section = request.args.get('company_section')

    if not screening_year or not company_section:
        return jsonify({'message': 'screening_year and company_section parameters are required'}), 400

    records = db.session.query(
        ScreeningBioData.id.label('record_id'),
        ScreeningBioData.patient_id_for_year.label('patient_id'),
        Patient.staff_id,
        Patient.first_name,
        Patient.last_name,
        Patient.department,
        Patient.gender,
        Patient.contact_phone
    ).join(
        Patient, ScreeningBioData.patient_comprehensive_id == Patient.id
    ).filter(
        ScreeningBioData.screening_year == screening_year,
        ScreeningBioData.company_section == company_section
    ).order_by(Patient.first_name, Patient.last_name).all()

    return jsonify([{
        'record_id': r.record_id,
        'patient_id': r.patient_id,
        'staff_id': r.staff_id,
        'first_name': r.first_name,
        'last_name': r.last_name,
        'department': r.department,
        'gender': r.gender,
        'contact_phone': r.contact_phone,
    } for r in records])

@bp.route('/screening/record/<int:record_id>', methods=['DELETE'])
@token_required('delete_screening_record')
def delete_screening_record(current_user, record_id):
    record = ScreeningBioData.query.get_or_404(record_id)
    log_audit(current_user, 'SCREENING_RECORD_DELETE', f"Deleted screening record for patient {record.patient_comprehensive.first_name} {record.patient_comprehensive.last_name} (Record ID: {record_id})")
    db.session.delete(record)
    db.session.commit()
    return jsonify({'message': 'Screening record deleted successfully.'}), 200

@bp.route('/consultations', methods=['POST'])
@token_required('perform_consultation')
def create_or_update_consultation(current_user):
    data = request.get_json()
    if not data or not data.get('staff_id'):
        return jsonify({'message': 'Staff ID is required'}), 400

    patient = Patient.query.filter_by(staff_id=data['staff_id']).first()
    if not patient:
        return jsonify({'message': 'Patient not found'}), 404

    consultation = Consultation.query.filter_by(patient_id=patient.id).first()
    if not consultation:
        consultation = Consultation(patient_id=patient.id)
        db.session.add(consultation)

    for key, value in data.items():
        if hasattr(consultation, key) and key != 'patient_id':
            setattr(consultation, key, value)

    log_audit(current_user, 'CONSULTATION_SAVE', f"Saved consultation for patient {patient.first_name} {patient.last_name} (Staff ID: {data['staff_id']})")
    db.session.commit()
    return jsonify({'message': 'Consultation saved successfully'}), 200

@bp.route('/consultations/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_consultation(current_user, staff_id):
    consultation = db.session.query(Consultation).join(Patient).filter(Patient.staff_id == staff_id).first()
    if not consultation:
        return jsonify({'message': 'Consultation not found'}), 404

    consultation_data = {key: getattr(consultation, key) for key in consultation.__table__.columns.keys()}
    return jsonify(consultation_data)

def create_or_update_test_result(model, staff_id):
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No input data provided'}), 400

    patient = Patient.query.filter_by(staff_id=staff_id).first()
    if not patient:
        return jsonify({'message': 'Patient not found'}), 404

    result = model.query.filter_by(patient_id=patient.id).first()
    if not result:
        result = model(patient_id=patient.id)
        db.session.add(result)

    for key, value in data.items():
        if hasattr(result, key) and key != 'patient_id':
            setattr(result, key, value)

    log_audit(current_user, 'TEST_RESULT_SAVE', f"Saved {model.__name__} for patient {patient.first_name} {patient.last_name} (Staff ID: {staff_id})")
    db.session.commit()
    return jsonify({'message': 'Test result saved successfully'}), 200

def get_test_result(model, staff_id):
    result = db.session.query(model).join(Patient).filter(Patient.staff_id == staff_id).first()
    if not result:
        return jsonify({'message': 'Test result not found'}), 404

    result_data = {key: getattr(result, key) for key in result.__table__.columns.keys()}
    return jsonify(result_data)

@bp.route('/test-results/full-blood-count/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_fbc(current_user, staff_id):
    return create_or_update_test_result(FullBloodCount, staff_id)

@bp.route('/test-results/full-blood-count/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_fbc(current_user, staff_id):
    return get_test_result(FullBloodCount, staff_id)

@bp.route('/test-results/kidney-function-test/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_kft(current_user, staff_id):
    return create_or_update_test_result(KidneyFunctionTest, staff_id)

@bp.route('/test-results/kidney-function-test/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_kft(current_user, staff_id):
    return get_test_result(KidneyFunctionTest, staff_id)

@bp.route('/test-results/lipid-profile/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_lp(current_user, staff_id):
    return create_or_update_test_result(LipidProfile, staff_id)

@bp.route('/test-results/lipid-profile/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_lp(current_user, staff_id):
    return get_test_result(LipidProfile, staff_id)

@bp.route('/test-results/liver-function-test/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_lft(current_user, staff_id):
    return create_or_update_test_result(LiverFunctionTest, staff_id)

@bp.route('/test-results/liver-function-test/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_lft(current_user, staff_id):
    return get_test_result(LiverFunctionTest, staff_id)

@bp.route('/test-results/ecg/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_ecg(current_user, staff_id):
    return create_or_update_test_result(ECG, staff_id)

@bp.route('/test-results/ecg/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_ecg(current_user, staff_id):
    return get_test_result(ECG, staff_id)

@bp.route('/test-results/spirometry/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_spirometry(current_user, staff_id):
    return create_or_update_test_result(Spirometry, staff_id)

@bp.route('/test-results/spirometry/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_spirometry(current_user, staff_id):
    return get_test_result(Spirometry, staff_id)

@bp.route('/test-results/audiometry/<string:staff_id>', methods=['POST'])
@token_required('enter_test_results')
def save_audiometry(current_user, staff_id):
    return create_or_update_test_result(Audiometry, staff_id)

@bp.route('/test-results/audiometry/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_audiometry(current_user, staff_id):
    return get_test_result(Audiometry, staff_id)

@bp.route('/screening/register', methods=['POST'])
@token_required('register_patient')
def register_for_screening(current_user):
    data = request.get_json()
    if not data:
        return jsonify({'message': 'No input data provided'}), 400

    # --- 1. Validate Input Data ---
    required_fields = [
        'staff_id', 'first_name', 'last_name', 'department', 'gender',
        'date_of_birth', 'contact_phone', 'email_address', 'race',
        'nationality', 'screening_year', 'company_section', 'patient_id_for_year'
    ]
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        return jsonify({'message': f'Missing required fields: {", ".join(missing)}'}), 400

    # --- 2. Find or Create Comprehensive Patient Record ---
    patient = Patient.query.filter_by(staff_id=data['staff_id']).first()
    dob = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()

    if patient:
        # Update existing patient's comprehensive data
        patient.first_name = data['first_name']
        patient.middle_name = data.get('middle_name')
        patient.last_name = data['last_name']
        patient.department = data['department']
        patient.gender = data['gender']
        patient.date_of_birth = dob
        patient.contact_phone = data['contact_phone']
        patient.email_address = data['email_address']
        patient.race = data['race']
        patient.nationality = data['nationality']
    else:
        # Create a new patient
        patient = Patient(
            staff_id=data['staff_id'],
            # The patient_id in the comprehensive table is not the yearly one
            patient_id=data['staff_id'], # Or generate another unique one
            first_name=data['first_name'],
            middle_name=data.get('middle_name'),
            last_name=data['last_name'],
            department=data['department'],
            gender=data['gender'],
            date_of_birth=dob,
            contact_phone=data['contact_phone'],
            email_address=data['email_address'],
            race=data['race'],
            nationality=data['nationality']
        )
        db.session.add(patient)

    # We need to commit here to get the patient.id if it's a new patient
    db.session.commit()

    # --- 3. Check for Duplicate Screening Registration ---
    existing_screening = ScreeningBioData.query.filter_by(
        patient_comprehensive_id=patient.id,
        screening_year=data['screening_year'],
        company_section=data['company_section']
    ).first()

    if existing_screening:
        return jsonify({'message': 'Patient already registered for this screening year and company section.'}), 409

    # Check if the year-specific patient ID is already in use for this context
    existing_patient_id_for_year = ScreeningBioData.query.filter_by(
        patient_id_for_year=data['patient_id_for_year'],
        screening_year=data['screening_year'],
        company_section=data['company_section']
    ).first()

    if existing_patient_id_for_year:
        return jsonify({'message': 'The patient ID for this year is already taken.'}), 409

    # --- 4. Create New ScreeningBioData Record ---
    new_screening_record = ScreeningBioData(
        patient_comprehensive_id=patient.id,
        patient_id_for_year=data['patient_id_for_year'],
        screening_year=data['screening_year'],
        company_section=data['company_section']
    )
    db.session.add(new_screening_record)
    log_audit(current_user, 'SCREENING_REGISTER', f"Registered patient {patient.first_name} {patient.last_name} for {data['screening_year']} screening.")
    db.session.commit()

    return jsonify({'message': 'Patient registered for screening successfully.'}), 201

@bp.route('/screening/stats', methods=['GET'])
@token_required('view_statistics')
def screening_stats(current_user):
    # --- 1. Get and Validate Query Parameters ---
    screening_year = request.args.get('screening_year', type=int)
    company_section = request.args.get('company_section')

    if not screening_year or not company_section:
        return jsonify({'message': 'screening_year and company_section parameters are required'}), 400

    # --- 2. Base Query for the given context ---
    base_query = ScreeningBioData.query.filter_by(
        screening_year=screening_year,
        company_section=company_section
    )

    # --- 3. Calculate Statistics ---
    total_registered = base_query.count()

    today_start = datetime.utcnow().date()
    registered_today = base_query.filter(func.date(ScreeningBioData.date_registered) == today_start).count()

    # Join with Patient table for gender and age
    query_with_join = base_query.join(Patient, ScreeningBioData.patient_comprehensive_id == Patient.id)

    male_count = query_with_join.filter(Patient.gender == 'Male').count()
    female_count = query_with_join.filter(Patient.gender == 'Female').count()

    # Age calculation (database-agnostic)
    # This method calculates the difference in years. It's simple, portable,
    # and sufficient for this statistical purpose.
    from sqlalchemy.sql import extract
    age_calc = extract('year', func.current_date()) - extract('year', Patient.date_of_birth)

    over_40_count = query_with_join.filter(age_calc >= 40).count()
    under_40_count = query_with_join.filter(age_calc < 40).count()

    # --- 4. Format and Return Response ---
    stats = {
        'total_registered': total_registered,
        'registered_today': registered_today,
        'male_count': male_count,
        'female_count': female_count,
        'over_40_count': over_40_count,
        'under_40_count': under_40_count
    }

    return jsonify(stats)

@bp.route('/screening/search', methods=['GET'])
@token_required('view_patient_data')
def search_screened_patients(current_user):
    # --- 1. Get and Validate Query Parameters ---
    screening_year = request.args.get('screening_year', type=int)
    company_section = request.args.get('company_section')
    search_term = request.args.get('searchTerm', '')

    if not screening_year or not company_section:
        return jsonify({'message': 'screening_year and company_section parameters are required'}), 400

    # --- 2. Build the Query ---
    query = db.session.query(
        Patient
    ).join(
        ScreeningBioData, Patient.id == ScreeningBioData.patient_comprehensive_id
    ).filter(
        ScreeningBioData.screening_year == screening_year,
        ScreeningBioData.company_section == company_section
    )

    if search_term:
        query = query.filter(Patient.staff_id.ilike(f'%{search_term}%'))

    # --- 3. Execute Query and Format Results ---
    patients = query.limit(10).all() # Limit results for performance

    results = [{
        'id': p.id,
        'staff_id': p.staff_id,
        'first_name': p.first_name,
        'last_name': p.last_name,
        'department': p.department,
        'age': (date.today().year - p.date_of_birth.year)
    } for p in patients]

    return jsonify(results)

@bp.route('/patient/<string:staff_id>', methods=['GET'])
@token_required('view_patient_data')
def get_patient_by_staff_id(current_user, staff_id):
    patient = Patient.query.filter_by(staff_id=staff_id).first_or_404()

    patient_data = {
        'id': patient.id,
        'staff_id': patient.staff_id,
        'first_name': patient.first_name,
        'last_name': patient.last_name,
        'department': patient.department,
        'age': (date.today().year - patient.date_of_birth.year)
    }
    return jsonify(patient_data)

# RBAC Routes
@bp.route('/users', methods=['GET'])
@token_required('manage_users')
def get_users(current_user):
    users = User.query.all()
    return jsonify([{'id': u.id, 'username': u.username, 'email': u.email, 'roles': [r.name for r in u.roles]} for u in users])

@bp.route('/roles', methods=['GET'])
@token_required('manage_roles')
def get_roles(current_user):
    roles = Role.query.all()
    return jsonify([{'id': r.id, 'name': r.name} for r in roles])

@bp.route('/user/<int:user_id>/assign-role', methods=['POST'])
@token_required('manage_users')
def assign_role(current_user, user_id):
    data = request.get_json()
    if not data or 'role_id' not in data:
        return jsonify({'message': 'Role ID is required'}), 400

    user = User.query.get_or_404(user_id)
    role = Role.query.get_or_404(data['role_id'])

    user.roles.append(role)
    log_audit(current_user, 'ROLE_ASSIGN', f"Assigned role '{role.name}' to user '{user.username}'")
    db.session.commit()

    return jsonify({'message': f'Role {role.name} assigned to user {user.username} successfully.'})

@bp.route('/permissions', methods=['GET'])
@token_required('manage_roles')
def get_permissions(current_user):
    permissions = Permission.query.all()
    return jsonify([{'id': p.id, 'name': p.name} for p in permissions])

@bp.route('/roles/<int:role_id>', methods=['GET'])
@token_required('manage_roles')
def get_role(current_user, role_id):
    role = Role.query.get_or_404(role_id)
    return jsonify({
        'id': role.id,
        'name': role.name,
        'permissions': [p.id for p in role.permissions]
    })

@bp.route('/roles', methods=['POST'])
@token_required('manage_roles')
def create_role(current_user):
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'message': 'Role name is required'}), 400

    new_role = Role(name=data['name'])
    db.session.add(new_role)
    log_audit(current_user, 'ROLE_CREATE', f"Created role '{new_role.name}'")
    db.session.commit()

    return jsonify({'id': new_role.id, 'name': new_role.name, 'permissions': []}), 201

@bp.route('/roles/<int:role_id>', methods=['PUT'])
@token_required('manage_roles')
def update_role(current_user, role_id):
    role = Role.query.get_or_404(role_id)
    data = request.get_json()

    if 'name' in data:
        role.name = data['name']

    if 'permission_ids' in data:
        role.permissions = []
        for p_id in data['permission_ids']:
            permission = Permission.query.get(p_id)
            if permission:
                role.permissions.append(permission)

    log_audit(current_user, 'ROLE_UPDATE', f"Updated role '{role.name}'")
    db.session.commit()
    return jsonify({'message': 'Role updated successfully.'})

@bp.route('/roles/<int:role_id>', methods=['DELETE'])
@token_required('manage_roles')
def delete_role(current_user, role_id):
    role = Role.query.get_or_404(role_id)

    # Basic protection for Admin role
    if role.name == 'Admin':
        return jsonify({'message': 'The Admin role cannot be deleted.'}), 403

    log_audit(current_user, 'ROLE_DELETE', f"Deleted role '{role.name}'")
    db.session.delete(role)
    db.session.commit()
    return jsonify({'message': 'Role deleted successfully.'})

# Audit Log Helper
def log_audit(user, action, details=""):
    """Helper function to create an audit log entry."""
    log_entry = AuditLog(user_id=user.id, action=action, details=details)
    db.session.add(log_entry)

# Temporary Access Code Routes
import uuid

@bp.route('/temp-codes', methods=['GET'])
@token_required('manage_roles')
def get_temp_codes(current_user):
    codes = TemporaryAccessCode.query.order_by(TemporaryAccessCode.id.desc()).all()
    return jsonify([{
        'id': c.id,
        'code': c.code,
        'permission': c.permission.name,
        'user': c.user.username if c.user else 'Any',
        'expiration': c.expiration.isoformat(),
        'use_type': c.use_type,
        'times_used': c.times_used,
        'is_active': c.is_active,
    } for c in codes])

@bp.route('/temp-codes', methods=['POST'])
@token_required('manage_roles')
def generate_temp_code(current_user):
    data = request.get_json()
    if not data or not data.get('permission_id') or not data.get('duration_minutes'):
        return jsonify({'message': 'Permission ID and duration are required'}), 400

    permission = Permission.query.get_or_404(data['permission_id'])
    duration = timedelta(minutes=int(data['duration_minutes']))
    expiration = datetime.utcnow() + duration
    code_str = f"LHCSL-{uuid.uuid4().hex[:8].upper()}"
    use_type = data.get('use_type', 'single-use')
    user_id = data.get('user_id') # Optional user ID

    new_code = TemporaryAccessCode(
        code=code_str,
        permission_id=permission.id,
        expiration=expiration,
        use_type=use_type,
        user_id=user_id if user_id else None
    )
    db.session.add(new_code)

    log_details = f"Generated code {code_str} for permission '{permission.name}'"
    if user_id:
        assigned_user = User.query.get(user_id)
        if assigned_user:
            log_details += f" for user '{assigned_user.username}'"
    log_audit(current_user, 'TEMP_CODE_GENERATED', log_details)
    db.session.commit()

    return jsonify({'message': 'Temporary access code generated successfully.', 'code': code_str}), 201

@bp.route('/temp-codes/activate', methods=['POST'])
@token_required()
def activate_temp_code(current_user):
    data = request.get_json()
    code_str = data.get('code')
    if not code_str:
        return jsonify({'message': 'Temporary access code is required'}), 400

    code = TemporaryAccessCode.query.filter_by(code=code_str, is_active=True).first()

    if not code or code.expiration < datetime.utcnow():
        log_audit(current_user, 'TEMP_CODE_ACTIVATE_FAILURE', f"Attempted to use invalid or expired code '{code_str}'")
        db.session.commit()
        return jsonify({'message': 'Invalid or expired code.'}), 404

    if code.user_id and code.user_id != current_user.id:
        log_audit(current_user, 'TEMP_CODE_ACTIVATE_FAILURE', f"User attempted to use code '{code_str}' assigned to another user.")
        db.session.commit()
        return jsonify({'message': 'This code is not assigned to you.'}), 403

    if code.use_type == 'single-use' and code.times_used > 0:
        log_audit(current_user, 'TEMP_CODE_ACTIVATE_FAILURE', f"Attempted to reuse single-use code '{code_str}'")
        db.session.commit()
        return jsonify({'message': 'This code has already been used.'}), 403

    # Grant permission in session (this is a simple way, a more robust system might use a separate table)
    # The decorator will need to be updated to check this session variable.
    session_permissions = session.get('temp_permissions', {})
    session_permissions[code.permission.name] = code.expiration.isoformat()
    session['temp_permissions'] = session_permissions

    code.times_used += 1
    if code.use_type == 'single-use':
        code.is_active = False

    log_audit(current_user, 'TEMP_CODE_ACTIVATE_SUCCESS', f"Successfully activated code '{code_str}' for permission '{code.permission.name}'")
    db.session.commit()

    return jsonify({'message': f"Permission '{code.permission.name}' granted temporarily."})

@bp.route('/temp-codes/<int:code_id>/revoke', methods=['POST'])
@token_required('manage_roles')
def revoke_temp_code(current_user, code_id):
    code = TemporaryAccessCode.query.get_or_404(code_id)
    code.is_active = False
    log_audit(current_user, 'TEMP_CODE_REVOKE', f"Revoked code '{code.code}'")
    db.session.commit()
    return jsonify({'message': 'Code revoked successfully.'})

@bp.route('/request-access', methods=['POST'])
@token_required()
def request_access(current_user):
    data = request.get_json()
    permission_needed = data.get('permission')
    if not permission_needed:
        return jsonify({'message': 'Permission is required'}), 400

    log_audit(current_user, 'TEMP_ACCESS_CODE_REQUEST', f"User requested access for permission: '{permission_needed}'")
    db.session.commit()

    return jsonify({'message': 'Your request for access has been logged for an administrator to review.'}), 200

@bp.route('/audit-logs', methods=['GET'])
@token_required('view_audit_log')
def get_audit_logs(current_user):
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).all()
    return jsonify([{
        'id': log.id,
        'username': log.user.username,
        'action': log.action,
        'timestamp': log.timestamp.isoformat(),
        'details': log.details,
    } for log in logs])


@bp.route('/user/<int:user_id>', methods=['GET'])
@token_required('view_users')
def get_user_details(current_user, user_id):
    user = User.query.get_or_404(user_id)
    return jsonify({
        'id': user.id,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'username': user.username,
        'email': user.email,
    })

@bp.route('/user/<int:user_id>', methods=['PUT'])
@token_required('edit_users')
def update_user_details(current_user, user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json()

    # Update standard fields
    user.first_name = data.get('first_name', user.first_name)
    user.last_name = data.get('last_name', user.last_name)
    user.username = data.get('username', user.username)
    user.email = data.get('email', user.email)

    # Update password if a new one is provided
    if 'new_password' in data and data['new_password']:
        if not User.is_password_strong(data['new_password']):
            return jsonify({'message': 'New password is not strong enough.'}), 400
        user.set_password(data['new_password'])
        log_audit(current_user, 'ADMIN_PASSWORD_CHANGE', f"Admin changed password for user '{user.username}'")

    log_audit(current_user, 'ADMIN_USER_UPDATE', f"Admin updated details for user '{user.username}'")
    db.session.commit()
    return jsonify({'message': 'User updated successfully.'})


@bp.route('/patient-report/email', methods=['POST'])
@token_required('email_report')
def email_patient_report(current_user):
    data = request.get_json()
    staff_id = data.get('staff_id')

    if not staff_id:
        return jsonify({"message": "Staff ID is required"}), 400

    patient = Patient.query.filter_by(staff_id=staff_id).first_or_404()

    # --- Email Logic ---
    sender_email_config = SystemConfig.query.filter_by(key='sender_email').first()
    app_password_config = SystemConfig.query.filter_by(key='app_password').first()

    if not sender_email_config or not app_password_config:
        return jsonify({'message': 'Email service is not configured.'}), 500

    sender_email = sender_email_config.value
    app_password = app_password_config.value

    # Get patient and screening details for the email body
    screening_record = ScreeningBioData.query.filter_by(patient_comprehensive_id=patient.id).order_by(ScreeningBioData.screening_year.desc()).first()

    if not screening_record:
        return jsonify({'message': 'No screening record found for this patient.'}), 404

    report_year = screening_record.screening_year
    organisation = screening_record.company_section

    # Create the email
    msg = EmailMessage()
    msg['Subject'] = f"MEDICAL REPORT: {report_year} Annual Medical Screening for SUNU Health Enrolees at {organisation} Obajana"
    msg['From'] = f"Legit HealthCare [{organisation}-OBAJANA] <{sender_email}>"
    msg['To'] = patient.email_address

    # Dynamic email body
    email_body = f"""
    Dear {patient.first_name} - {patient.staff_id},

    Thanks for making yourself available for this year's Annual Medical Screening Exercise.

    Please find attached your Medical Report for the year {report_year}.
    The hard copy of your result will be given to the {organisation} Medical Team for dispatch to your department.

    Warm Regards

    SIGNED:
    Dr. Anyanwu Ugochukwu D. FMCPath
    Consultant in Charge
    Legit HealthCare Services Ltd.

    <br>
    For more info and complaints, contact the Consultant in charge (Doctor) on WhatsApp at: https://bit.ly/SUNU-Doctor or Systems Administrator at: https://bit.ly/Admin_lhcsl
    """
    msg.set_content(email_body, subtype='html')

    # TODO: Add PDF attachment logic here in a future step.
    # For now, just sending the text email.

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(sender_email, app_password)
            smtp.send_message(msg)

        log_audit(current_user, 'EMAIL_REPORT_SUCCESS', f"Report successfully emailed to {patient.email_address} for staff ID {staff_id}")
        db.session.commit()
        return jsonify({"message": "Report sent successfully"}), 200
    except Exception as e:
        current_app.logger.error(f"Email failed to send: {e}")
        log_audit(current_user, 'EMAIL_REPORT_FAILURE', f"Failed to email report to {patient.email_address} for staff ID {staff_id}: {e}")
        db.session.commit()
        return jsonify({"message": "Failed to send email"}), 500


# System Config Routes
@bp.route('/config/email', methods=['GET'])
@token_required('manage_roles') # Or a new 'manage_system_config' permission
def get_email_config(current_user):
    sender_email = SystemConfig.query.filter_by(key='sender_email').first()
    return jsonify({
        'sender_email': sender_email.value if sender_email else ''
    })

@bp.route('/config/email', methods=['POST'])
@token_required('manage_roles')
def set_email_config(current_user):
    data = request.get_json()

    sender_email = data.get('sender_email')
    app_password = data.get('app_password')

    if not sender_email or not app_password:
        return jsonify({'message': 'Sender email and app password are required.'}), 400

    # --- Update or Create Sender Email ---
    email_config = SystemConfig.query.filter_by(key='sender_email').first()
    if not email_config:
        email_config = SystemConfig(key='sender_email', value=sender_email)
        db.session.add(email_config)
    else:
        email_config.value = sender_email

    # --- Update or Create App Password ---
    password_config = SystemConfig.query.filter_by(key='app_password').first()
    if not password_config:
        password_config = SystemConfig(key='app_password', value=app_password)
        db.session.add(password_config)
    else:
        password_config.value = app_password

    log_audit(current_user, 'CONFIG_UPDATE', 'Updated email configuration.')
    db.session.commit()
    return jsonify({'message': 'Email configuration saved successfully.'})


# --- Messaging Routes ---
@bp.route('/conversations', methods=['GET'])
@token_required()
def get_conversations(current_user):
    conversations = current_user.conversations
    return jsonify([{
        'id': c.id,
        'name': c.name or ", ".join([p.username for p in c.participants if p.id != current_user.id]),
        'is_group': c.is_group,
        'last_message': c.messages.order_by(Message.timestamp.desc()).first().content if c.messages.first() else None
    } for c in conversations])

@bp.route('/conversations/<int:conversation_id>', methods=['GET'])
@token_required()
def get_messages(current_user, conversation_id):
    conversation = Conversation.query.get_or_404(conversation_id)
    if current_user not in conversation.participants:
        return jsonify({'message': 'Not a participant of this conversation'}), 403

    messages = conversation.messages.order_by(Message.timestamp.asc()).all()
    return jsonify([{
        'id': m.id,
        'sender_id': m.sender_id,
        'sender_username': m.sender.username,
        'content': m.content,
        'timestamp': m.timestamp.isoformat()
    } for m in messages])

@bp.route('/messages', methods=['POST'])
@token_required()
def send_message(current_user):
    data = request.get_json()
    conversation_id = data.get('conversation_id')
    content = data.get('content')
    recipient_id = data.get('recipient_id') # For starting a new 1-on-1 chat

    if not content:
        return jsonify({'message': 'Message content is required'}), 400

    if conversation_id:
        conversation = Conversation.query.get_or_404(conversation_id)
        if current_user not in conversation.participants:
            return jsonify({'message': 'Not a participant of this conversation'}), 403
    elif recipient_id:
        recipient = User.query.get(recipient_id)
        if not recipient:
            return jsonify({'message': 'Recipient not found'}), 404

        # Check if a 1-on-1 conversation already exists
        conversation = Conversation.query.filter(Conversation.is_group == False) \
            .filter(Conversation.participants.contains(current_user)) \
            .filter(Conversation.participants.contains(recipient)).first()

        if not conversation:
            conversation = Conversation(participants=[current_user, recipient])
            db.session.add(conversation)
    else:
        return jsonify({'message': 'conversation_id or recipient_id is required'}), 400

    new_message = Message(
        conversation_id=conversation.id,
        sender_id=current_user.id,
        content=content
    )
    db.session.add(new_message)
    db.session.commit()
    return jsonify({'message': 'Message sent successfully.'}), 201

# User Self-Service Routes
@bp.route('/profile', methods=['PUT'])
@token_required()
def update_profile(current_user):
    data = request.get_json()
    # Check password for security
    if not current_user.check_password(data.get('current_password', '')):
        return jsonify({'message': 'Incorrect password'}), 403

    current_user.first_name = data.get('first_name', current_user.first_name)
    current_user.last_name = data.get('last_name', current_user.last_name)
    current_user.username = data.get('username', current_user.username)
    current_user.email = data.get('email', current_user.email)

    db.session.commit()
    log_audit(current_user, 'PROFILE_UPDATE', 'User updated their own profile.')
    db.session.commit()
    return jsonify({'message': 'Profile updated successfully.'})

@bp.route('/profile/change-password', methods=['POST'])
@token_required()
def change_password(current_user):
    data = request.get_json()
    if not current_user.check_password(data.get('current_password')):
        return jsonify({'message': 'Incorrect current password'}), 403

    new_password = data.get('new_password')
    if not User.is_password_strong(new_password):
        return jsonify({'message': 'New password is not strong enough.'}), 400

    current_user.set_password(new_password)
    db.session.commit()
    log_audit(current_user, 'PASSWORD_CHANGE', 'User changed their own password.')
    db.session.commit()
    return jsonify({'message': 'Password updated successfully.'})

# 2FA Routes
@bp.route('/2fa/status', methods=['GET'])
@token_required()
def get_2fa_status(current_user):
    return jsonify({'enabled': current_user.otp_enabled})

@bp.route('/2fa/enable', methods=['POST'])
@token_required()
def enable_2fa(current_user):
    if current_user.otp_enabled:
        return jsonify({'message': '2FA is already enabled.'}), 400

    # Generate a new secret
    current_user.otp_secret = pyotp.random_base32()
    db.session.commit()

    # Generate QR code
    totp = pyotp.TOTP(current_user.otp_secret)
    provisioning_uri = totp.provisioning_uri(name=current_user.email, issuer_name='Legit HealthCare')

    img = qrcode.make(provisioning_uri)
    buf = io.BytesIO()
    img.save(buf)
    buf.seek(0)
    img_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')

    return jsonify({'qr_code_url': f'data:image/png;base64,{img_b64}'})

@bp.route('/2fa/verify', methods=['POST'])
@token_required()
def verify_2fa(current_user):
    data = request.get_json()
    otp = data.get('otp')

    totp = pyotp.TOTP(current_user.otp_secret)
    if not totp.verify(otp):
        return jsonify({'message': 'Invalid OTP'}), 400

    current_user.otp_enabled = True
    db.session.commit()
    log_audit(current_user, '2FA_ENABLED', 'User enabled 2FA.')
    db.session.commit()
    return jsonify({'message': '2FA enabled successfully.'})

@bp.route('/2fa/disable', methods=['POST'])
@token_required()
def disable_2fa(current_user):
    current_user.otp_enabled = False
    db.session.commit()
    log_audit(current_user, '2FA_DISABLED', 'User disabled 2FA.')
    db.session.commit()
    return jsonify({'message': '2FA disabled successfully.'})

# Download Routes
@bp.route('/patients/download', methods=['GET'])
@token_required('download_patient_biodata')
def download_all_patients(current_user):
    patients = Patient.query.order_by(Patient.first_name, Patient.last_name).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'All Patients Bio-Data'

    headers = [
        'Staff ID', 'First Name', 'Middle Name', 'Last Name', 'Department',
        'Gender', 'Date of Birth', 'Contact Phone', 'Email Address',
        'Race', 'Nationality'
    ]
    sheet.append(headers)

    for p in patients:
        sheet.append([
            p.staff_id, p.first_name, p.middle_name, p.last_name, p.department,
            p.gender, p.date_of_birth.isoformat() if p.date_of_birth else '', p.contact_phone,
            p.email_address, p.race, p.nationality
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    log_audit(current_user, 'DOWNLOAD_PATIENT_BIODATA', 'Downloaded comprehensive patient bio-data.')
    db.session.commit()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='all_patients_biodata.xlsx'
    )

@bp.route('/screening/download', methods=['GET'])
@token_required('download_screening_data')
def download_screening_data(current_user):
    screening_year = request.args.get('screening_year', type=int)
    company_section = request.args.get('company_section')

    if not screening_year or not company_section:
        return jsonify({'message': 'screening_year and company_section parameters are required'}), 400

    records = db.session.query(
        Patient
    ).join(
        ScreeningBioData, Patient.id == ScreeningBioData.patient_comprehensive_id
    ).filter(
        ScreeningBioData.screening_year == screening_year,
        ScreeningBioData.company_section == company_section
    ).order_by(Patient.first_name, Patient.last_name).all()

    def get_patient_summary_for_download(patient):
        summary = {
            'staff_id': patient.staff_id,
            'first_name': patient.first_name,
            'middle_name': patient.middle_name,
            'last_name': patient.last_name,
            'department': patient.department,
            'gender': patient.gender,
            'date_of_birth': patient.date_of_birth.isoformat() if patient.date_of_birth else '',
        }

        def model_to_dict(model_instance):
            if not model_instance:
                return {}
            exclude = ['id', 'patient_id']
            result = {}
            for c in model_instance.__table__.columns:
                if c.name not in exclude:
                    value = getattr(model_instance, c.name)
                    if isinstance(value, (datetime, date)):
                        result[c.name] = value.isoformat()
                    else:
                        result[c.name] = value
            return result

        summary.update(model_to_dict(patient.consultation))
        summary.update(model_to_dict(patient.full_blood_count))
        summary.update(model_to_dict(patient.kidney_function_test))
        summary.update(model_to_dict(patient.lipid_profile))
        summary.update(model_to_dict(patient.liver_function_test))
        summary.update(model_to_dict(patient.ecg))
        summary.update(model_to_dict(patient.spirometry))
        summary.update(model_to_dict(patient.audiometry))
        return summary

    patient_summaries = [get_patient_summary_for_download(p) for p in records]

    if not patient_summaries:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'No Records'
        sheet.append(['No records found for the selected criteria.'])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'screening_data_{screening_year}_{company_section}.xlsx'
        )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'{screening_year} {company_section} Screening'

    headers = list(patient_summaries[0].keys())
    sheet.append(headers)

    for summary in patient_summaries:
        row = [summary.get(h) for h in headers]
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    log_audit(current_user, 'DOWNLOAD_SCREENING_DATA', f'Downloaded screening data for {screening_year} {company_section}.')
    db.session.commit()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'screening_data_{screening_year}_{company_section}.xlsx'
    )

@bp.route('/screening/biodata/download', methods=['GET'])
@token_required('download_screening_biodata')
def download_screening_biodata(current_user):
    screening_year = request.args.get('screening_year', type=int)
    company_section = request.args.get('company_section')

    if not screening_year or not company_section:
        return jsonify({'message': 'screening_year and company_section parameters are required'}), 400

    records = db.session.query(
        Patient
    ).join(
        ScreeningBioData, Patient.id == ScreeningBioData.patient_comprehensive_id
    ).filter(
        ScreeningBioData.screening_year == screening_year,
        ScreeningBioData.company_section == company_section
    ).order_by(Patient.first_name, Patient.last_name).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'{screening_year} {company_section} Bio-Data'

    headers = [
        'Staff ID', 'First Name', 'Middle Name', 'Last Name', 'Department',
        'Gender', 'Date of Birth', 'Contact Phone', 'Email Address',
        'Race', 'Nationality'
    ]
    sheet.append(headers)

    for p in records:
        sheet.append([
            p.staff_id, p.first_name, p.middle_name, p.last_name, p.department,
            p.gender, p.date_of_birth.isoformat() if p.date_of_birth else '', p.contact_phone,
            p.email_address, p.race, p.nationality
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    log_audit(current_user, 'DOWNLOAD_SCREENING_BIODATA', f'Downloaded screening bio-data for {screening_year} {company_section}.')
    db.session.commit()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'screening_biodata_{screening_year}_{company_section}.xlsx'
    )


# Branding Routes
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/branding', methods=['GET'])
def get_branding():
    branding = Branding.query.first()
    if not branding:
        # Create a default entry if it doesn't exist
        branding = Branding()
        db.session.add(branding)
        db.session.commit()

    return jsonify({
        'clinic_name': branding.clinic_name,
        'logo_light': branding.logo_light,
        'logo_dark': branding.logo_dark,
        'logo_home': branding.logo_home,
        'report_header': branding.report_header,
        'report_signature': branding.report_signature,
        'report_footer': branding.report_footer,
        'doctor_name': branding.doctor_name,
        'doctor_title': branding.doctor_title,
    })

@bp.route('/branding', methods=['POST'])
@token_required('manage_branding')
def update_branding(current_user):
    branding = Branding.query.first()
    if not branding:
        branding = Branding()
        db.session.add(branding)

    # --- Update Clinic Name ---
    if 'clinic_name' in request.form:
        branding.clinic_name = request.form['clinic_name']
    if 'doctor_name' in request.form:
        branding.doctor_name = request.form['doctor_name']
    if 'doctor_title' in request.form:
        branding.doctor_title = request.form['doctor_title']

    # --- Handle File Uploads ---
    files = request.files
    for key in ['logo_light', 'logo_dark', 'logo_home', 'report_header', 'report_signature', 'report_footer']:
        if key in files:
            file = files[key]
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # To avoid filename collisions, prepend with a unique identifier
                unique_filename = f"{key}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
                save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(save_path)
                setattr(branding, key, unique_filename) # Store the filename in the DB

    log_audit(current_user, 'BRANDING_UPDATE', 'Updated branding settings.')
    db.session.commit()

    return jsonify({'message': 'Branding updated successfully.'})

@bp.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@bp.route('/patients/claim-info', methods=['GET'])
def get_patient_claim_info():
    staff_id = request.args.get('staff_id')
    if not staff_id:
        return jsonify({'message': 'Staff ID is required'}), 400

    patient = Patient.query.filter_by(staff_id=staff_id).first()
    if not patient:
        return jsonify({'message': 'Patient not found'}), 404

    if patient.user_id:
        return jsonify({'message': 'Account already claimed'}), 409

    return jsonify({
        'first_name': patient.first_name,
        'last_name': patient.last_name,
        'email': patient.email_address,
    })

@bp.route('/patients/upload', methods=['POST'])
@token_required('upload_patient_data')
def upload_patients(current_user):
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    if file and file.filename.endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Assuming the first row is the header
            headers = [cell.value for cell in sheet[1]]

            required_fields = ['Staff ID', 'First Name', 'Last Name']
            if not all(field in headers for field in required_fields):
                return jsonify({'message': 'Missing required columns: Staff ID, First Name, Last Name'}), 400

            patients_added = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))

                staff_id = row_data.get('Staff ID')
                first_name = row_data.get('First Name')
                last_name = row_data.get('Last Name')

                if not staff_id or not first_name or not last_name:
                    continue # Skip rows with missing required data

                if Patient.query.filter_by(staff_id=staff_id).first():
                    continue # Skip existing patients

                new_patient = Patient(
                    staff_id=staff_id,
                    first_name=first_name,
                    last_name=last_name,
                    department=row_data.get('Department'),
                    gender=row_data.get('Gender'),
                    date_of_birth=row_data.get('Date of Birth'),
                    contact_phone=row_data.get('Contact Phone'),
                    email_address=row_data.get('Email Address'),
                    race=row_data.get('Race'),
                    nationality=row_data.get('Nationality')
                )
                db.session.add(new_patient)
                patients_added += 1

            db.session.commit()
            log_audit(current_user, 'PATIENT_BULK_UPLOAD', f'Uploaded and added {patients_added} new patients.')
            db.session.commit()

            return jsonify({'message': f'Successfully added {patients_added} new patients.'}), 201

        except Exception as e:
            current_app.logger.error(f"Error processing patient upload file: {e}")
            return jsonify({'message': 'Error processing file'}), 500

    return jsonify({'message': 'Invalid file type'}), 400